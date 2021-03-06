from openpyxl import load_workbook
from root.nested import get_logger
from root.nested.visualize.extend_bokeh import ExtendBokeh
import pandas as pd
import requests

class CboExcelApi:

    def __init__(self,
                 xlsx_file_name):

        self.logger = get_logger()
        self.workbook = load_workbook(xlsx_file_name)

    def downloadCboData(self):

        budget_economic_data_url = 'https://www.cbo.gov/about/products/budget-economic-data'
        historical_budget_data_url = 'https://www.cbo.gov/system/files/2019-01/51134-2019-01-historicalbudgetdata.xlsx'
        r = requests.get(historical_budget_data_url)
        with open('/Users/traderghazy/workspace/data/cbo/51134-2019-01-historicalbudgetdata.xlsx', 'wb') as f:
            f.write(r.content)
        self.logger.info("CboExcelApi.downloadCbodata(): %s ", r.status_code)
        self.logger.info("CboExcelApi.downloadCboData(): %s ", r.headers['content_type'])
        self.logger.info("CboExcelApi.downloadCboData(): %s ", r.encoding)

        return

    def doWork(self):

        sheetnames = self.workbook.get_sheet_names()
        self.logger.info("the sheet names in the input xlsx file are %s", str(sheetnames))
        revenues_sheet = self.workbook.get_sheet_by_name('2. Revenues')
        outlays_sheet = self.workbook.get_sheet_by_name('3. Outlays')
        begin_tabledata_row_num = 0
        title_of_revenues_plot = ""
        for cell_obj in revenues_sheet['A1':'A200']:
            for cell in cell_obj:
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith("2. Revenues"):
                        title_of_revenues_plot = cell.value
                        begin_tabledata_row_num = cell.row + 2
        # row_num now set to where table data starts
        # check first to make sure its not empty
        table_header_row = revenues_sheet[begin_tabledata_row_num]
        clean_header_list = []
        for header_cell in table_header_row:
            if header_cell.value is not None:
                clean_header_list.append(header_cell.value)
        # now, look for the next empty row to know where the table ends
        revenue_data_dict = {}
        revenue_perc_gdp_data_dict = {}
        for row in revenues_sheet[begin_tabledata_row_num+1:200]:
            # ignore any row with out a year value (which is first column)
            if isinstance(row[0].value, str) and row[0].value.startswith("Sources:"):
                # reached end of sheet.
                break
            if row[0].value is not None:# and row[0].value != "Sources: Congressional Budget Office; Office of Management and Budget.":							:
                val_list=[]
                for row_val in row:
                    if row_val.value is not None:
                        val_list.append(row_val.value)
                if (row[0].value in revenue_data_dict):
                    # we are done with the revenues ($) table, % of GDP table being read now.
                    revenue_perc_gdp_data_dict[row[0].value] = val_list
                else:
                    revenue_data_dict[row[0].value] = val_list
        clean_header_list.insert(0, "Year")
        revenues_df = pd.DataFrame.from_dict(revenue_data_dict,
                                             orient='index')
                                             #columns = clean_header_list)
        revenues_perc_gdp__df = pd.DataFrame.from_dict(revenue_perc_gdp_data_dict,
                                                      orient='index')
        revenues_perc_gdp__df.columns = clean_header_list
        revenues_df.columns = clean_header_list
        return revenues_df, revenues_perc_gdp__df, title_of_revenues_plot, title_of_revenues_plot + ", As % of GDP"

    def visualizeCbo(self,
                     df,
                     title,
                     xlabel,
                     ylabel,
                     html_output_filename):

        ExtendBokeh.visualizeCbo(df, title, xlabel, ylabel, html_output_filename)

if __name__ == "__main__":

    excel_filename = "/Users/traderghazy/workspace/data/cbo/51134-2019-01-historicalbudgetdata.xlsx"
    cboeapi = CboExcelApi(excel_filename)
    revenues_df, revenues_perc_gdp_df, title_of_revenues_plot, title_of_revenues_plot_aspctGdp = cboeapi.doWork()
    revenues_output_filename = "/Users/traderghazy/workspace/data/bokeh/html/CBO_revenues.html"
    revenues_aspctGdp_output_filename = "/Users/traderghazy/workspace/data/bokeh/html/CBO_revenues_pctGdp.html"
    cboeapi.visualizeCbo(df=revenues_df,
                         title=title_of_revenues_plot,
                         xlabel="Year",
                         ylabel="In Billions (US $)",
                         html_output_filename=revenues_output_filename)
    cboeapi.visualizeCbo(df=revenues_perc_gdp_df,
                         title=title_of_revenues_plot_aspctGdp,
                         xlabel="Year",
                         ylabel="% of GDP",
                         html_output_filename=revenues_aspctGdp_output_filename)